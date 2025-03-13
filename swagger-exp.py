import json
import re
from requests.exceptions import JSONDecodeError
from openpyxl import Workbook,load_workbook
import requests
import argparse
import base64
from colorama import init,Fore,Style

import warnings
warnings.filterwarnings('ignore')

send_data_list = []

def args():
    parser = argparse.ArgumentParser()
    parser.add_argument('-u','--url',dest="url",required=True, type=str,help="指定测试的URL(e.g. -u http://127.0.0.1/v2/api-docs)")
    parser.add_argument('-p', '--proxy', dest="proxy", type=str,help="设置代理(e.g. -p http://127.0.0.1:8080)")
    parser.add_argument('-path', '--path', dest="path", default='', type=str,
                        help="指定路径，默认api是根路径拼接，如http://xxx.xx/user/api，但是实际路径可能是http://xxx.xx/admin/user/api，那么这个admin就需要该参数来指定。")
    parser.add_argument('-cookie', '--cookie', dest="cookie", default='', type=str, help="指定cookie")
    parser.add_argument('-m', '--mode', dest="mode", default='sec', type=str,
                        help="指定模式，默认是sec（只测试查询接口），也可以指定为all测试所有接口。")

    return parser.parse_args()

def print_raw(raw,url,type,host=""):
    print_str = ""
    if type=="req":
        # print('>>>>>请求包>>>>>')
        print_str += raw['method'] + " " + url + " HTTP/1.1\n"
        print_str += "Host: " + host + "\n"
        headers = raw['headers']
        for key,values in headers.items():
            print_str += key + ": " + values + "\n"
        if raw['body'] != None:
            print_str += "\n" + str(raw['body'])
        else:
            print_str += "\n"
    elif type=="rep":
        # print('<<<<<响应包<<<<<')
        print_str += "HTTP/1.1 " + str(raw['status_code']) + "\n"
        headers = raw['headers']
        for key, values in headers.items():
            print_str += key + ": " + values + "\n"
        # print(raw['_content'])
        try:
            print_str += "\n" + raw['_content'].decode('utf-8')
        except Exception as e:
            print(Style.BRIGHT + Fore.RED + "[-] 响应包内容可能为二进制文件数据，因此做了base64编码输出：")
            base64_str = base64.b64encode(raw['_content']).decode('utf-8')
            print_str += "\n" + "响应包内容可能为二进制文件数据，因此做了base64编码输出：\n" + base64_str

    # print(print_str)
    return print_str

def print_api(data):
    tmp_data = []
    for api_data in data:
        # print("========================\n接口：" + api_data[0])
        # print("描述：" + api_data[2])
        # print("请求包：")
        ret = re.search('^http[s]?://(?P<host>.+\.[0-9a-zA-Z-]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?)[/]', api_data[0])
        host = ret.group('host')
        print_str = api_data[4].upper() + " " + '/'+re.sub('^http[s]?://.+\.[0-9a-zA-Z-]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?[/]','',api_data[0]) + " HTTP/1.1\n"
        print_str += "Host: " + host + "\n"
        for key, values in api_data[1].items():
            print_str += key + ": " + values + "\n"
        print_str += "\n" + api_data[3]
        tmp_data.append([api_data[0],api_data[2],print_str])
        # print(print_str)
    return tmp_data

def Scanner(url,headers,method,proxies,summary,data=""):
    try:
        if method == "get":
            # print(headers)
            rep = requests.get(url,headers=headers,verify=False,proxies=proxies,allow_redirects=False)
        elif method == "options":
            rep = requests.options(url, headers=headers, verify=False, data=data, proxies=proxies,
                                   allow_redirects=False)
        elif method == "put" and (args.mode == "all"):
            rep = requests.put(url, headers=headers, verify=False, data=data, proxies=proxies, allow_redirects=False)
        elif method == "post":
            rep = requests.post(url, headers=headers, verify=False, data=data, proxies=proxies, allow_redirects=False)
        else:
            print(Style.BRIGHT + Fore.RED + "[-] 暂不支持的请求类型！请求类型为：" + method)
            return False
    except Exception as e:
        print(Style.BRIGHT + Fore.RED + "[-] 请求错误！请求URL："+url+" ，报错信息：" + str(e))
        send_data_list.append([url, summary, "请求错误！", "请求错误！", "请求错误！请求URL："+url+" ，报错信息：" + str(e), "请求错误！请求URL："+url+" ，报错信息：" + str(e)])
        return False

    print(Style.BRIGHT + Fore.YELLOW + '[*] ' + url + "  [code:" + str(rep.status_code) + "]" + "   [size:" + str(len(rep.text)) + "]")
    ret = re.search('^http[s]?://(?P<host>.+\.[0-9a-zA-Z-]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?)[/]', url)
    host = ret.group('host')
    req_data = print_raw(rep.request.__dict__,'/'+re.sub('^http[s]?://.+\.[0-9a-zA-Z]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?[/]','',url),type='req',host=host)
    rep_data = print_raw(rep.__dict__, '/'+re.sub('^http[s]?://.+\.[0-9a-zA-Z]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?[/]','',url), type='rep')
    send_data_list.append([url,summary,str(rep.status_code),str(len(rep.text)),req_data,rep_data])

def screen(path):
    str = path.lower()
    str_list = str.split('/')
    flag = []
    for i in str_list:
        if i.startswith('get') or i.startswith('query') or i.startswith('select') or i.startswith('search') or i.startswith('show') or i.startswith('list') or (i=="info"):
            flag.append('get')
        elif i.find('upload') != -1 or i.find('fileupload') != -1or i.find('uploadfile') != -1 or i.find('uploads') != -1:
            flag.append('upload')
        elif i.find('download') != -1 or i.find('filedownload') != -1 or i.find('downloadfile') != -1 or i.find('downloads') != -1:
            flag.append('download')
        elif i.find('adduser') != -1 or i.find('useradd') != -1 or i.find('rest-pwd') != -1 or i.find('register') != -1 or i.find('password') != -1 or i.find('updatepwd') != -1 or i.find('changepwd') != -1 or (i=="user"):
            flag.append('user')
    return flag

def get_definitions(data,definition,method):
    if method == "post":
        parameters = {}
        if "#/definitions/" in definition:
            definition = definition.replace('#/definitions/','')
            for i in data['definitions']:
                if i == definition:
                    for parameter in data['definitions'][i]['properties']:
                        # print(parameter)
                        try:
                            ref = data['definitions'][i]['properties'][parameter]['$ref']
                            parameters[parameter] = get_definitions(data,ref,'post')
                        except Exception as e:
                            if data['definitions'][i]['properties'][parameter]['type'] == "integer":
                                parameters[parameter] = 1
                            elif data['definitions'][i]['properties'][parameter]['type'] == 'number':
                                parameters[parameter] = 2.0
                            elif data['definitions'][i]['properties'][parameter]['type'] == 'array':
                                parameters[parameter] = '[]'
                            else:
                                parameters[parameter] = "string"
        elif "#/components/schemas/" in definition:
            definition = definition.replace('#/components/schemas/','')
            for i in data['components']['schemas']:
                if i == definition:
                    for parameter in data['components']['schemas'][i]['properties']:
                        # print(parameter)
                        try:
                            test1 = data['components']['schemas'][i]['properties'][parameter]
                            ref1 = test1[list(test1.keys())[0]][0]['$ref']
                            # print(ref1)
                            parameters[parameter] = get_definitions(data, ref1, 'post')
                        except Exception as ee:
                            # print(ee)
                            try:
                                ref = data['components']['schemas'][i]['properties'][parameter]['$ref']
                                parameters[parameter] = get_definitions(data,ref,'post')
                            except Exception as e:
                                try:
                                    if data['components']['schemas'][i]['properties'][parameter]['type'] == "integer":
                                        parameters[parameter] = 1
                                    elif data['components']['schemas'][i]['properties'][parameter]['type'] == 'number':
                                        parameters[parameter] = 2.0
                                    elif data['components']['schemas'][i]['properties'][parameter]['type'] == 'array':
                                        parameters[parameter] = '[]'
                                    else:
                                        parameters[parameter] = "string"
                                except Exception as type_e:
                                    parameters[parameter] = "string"
    elif method == "get":
        parameters = []
        if "#/definitions/" in definition:
            definition = definition.replace('#/definitions/','')
            for i in data['definitions']:
                if i == definition:
                    for parameter in data['definitions'][i]['properties']:
                        if data['definitions'][i]['properties'][parameter]['type'] == "integer":
                            parameters.append(parameter + "=1")
                        elif data['definitions'][i]['properties'][parameter]['type'] == 'number':
                            parameters.append(parameter + "=2.0")
                        elif data['definitions'][i]['properties'][parameter]['type'] == 'array':
                            parameters.append(parameter + "=[]")
                        else:
                            parameters.append(parameter + "=string")
        elif method == "get":
            definition = definition.replace('#/components/schemas/','')
            for i in data['components']['schemas']:
                if i == definition:
                    for parameter in data['components']['schemas'][i]['properties']:
                        if data['components']['schemas'][i]['properties'][parameter]['type'] == "integer":
                            parameters.append(parameter + "=1")
                        elif data['components']['schemas'][i]['properties'][parameter]['type'] == 'number':
                            parameters.append(parameter + "=2.0")
                        elif data['components']['schemas'][i]['properties'][parameter]['type'] == 'array':
                            parameters.append(parameter + "=[]")
                        else:
                            parameters.append(parameter + "=string")
    return parameters

def get_method(rep,path,method,url1,cookie):
    try:
        content_type = rep['paths'][path][method]['consumes'][0]
    except Exception as e:
        content_type = "text/html; charset=utf-8"
    parameters = []
    if cookie != '':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'Content-Type': content_type,
            'Cookie': cookie
        }
    else:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'Content-Type': content_type
        }
    try:
        summary = rep['paths'][path][method]['summary']
    except Exception as e:
        summary = ""
    try:
        if re.search('\{.*?\}', path):
            # parameter = re.sub('\{.*?\}', '1', path)
            # parameter = re.sub('\{', '', path)
            # parameter = re.sub('\}', '', parameter)
            new_url = url1 + path
        else:
            try:
                is_parameters = rep['paths'][path][method]['parameters']
                try:
                    definition = rep['paths'][path][method]['parameters'][0]['schema']['$ref']
                except Exception as e:
                    definition = ""
                if definition != "":
                    parameters = get_definitions(rep, definition, 'get')
                else:
                    for parameter in rep['paths'][path][method]['parameters']:
                        try:
                            default_str = parameter['default'] # 获取默认值
                        except Exception as e:
                            # print(parameter['type'])
                            try:
                                type = parameter['schema']['type'] # 3.0版本
                            except Exception as e:
                                try:
                                    type = parameter['type'] # 2.0和1.0版本
                                except:
                                    type = 'string'
                            if type == "integer":
                                default_str = 1
                            elif type == "array":
                                default_str = '[]'
                            else:
                                default_str = "string"
                        if  parameter['in'] == "header": # 参数位置在header
                            headers[parameter['name']] = str(default_str)
                        else:
                            parameters.append(parameter['name'] + "=" + str(default_str))
                new_url = url1 + path + '?' + '&'.join(parameters)
            except Exception as e:
                # print(e)
                new_url = url1 + path
        return new_url,headers,summary,"",method
    except Exception as e:
        return False

def post_method(rep,path,method,url1,cookie):
    try:
        content_type = rep['paths'][path][method]['consumes'][0]
    except Exception as e:
        content_type = "text/html; charset=utf-8"
    parameters = {}
    if cookie != '':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'Content-Type': content_type,
            'Cookie': cookie
        }
    else:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'Content-Type': content_type
        }
    try:
        summary = rep['paths'][path][method]['summary']
    except Exception as e:
        summary = ""
    parameters1 = {}
    try:
        # 3.0的另一种形式
        try:
            content = rep['paths'][path][method]['requestBody']['content']
            headers['Content-Type'] = list(content.keys())[0] # 获取第一个key
            definition1 = rep['paths'][path][method]['requestBody']['content'][list(content.keys())[0]]['schema']['$ref']
            parameters1 = get_definitions(rep, definition1, 'post')
        except Exception as e:
            # print(e)
            pass
        try:
            definition = rep['paths'][path][method]['parameters'][0]['schema']['$ref']
        except Exception as e:
            definition = ""
        if definition != "":
            parameters = get_definitions(rep, definition, 'post')
        else:
            # 如果没有任何参数，报错，则忽略。
            try:
                for parameter in rep['paths'][path][method]['parameters']:
                    try:
                        default_str = parameter['default']
                    except Exception as e:
                        try:
                            type = parameter['schema']['type'] # 3.0版本
                        except Exception as e:
                            try:
                                type = parameter['type'] # 2.0和1.0版本
                            except Exception as e:
                                type = "string"
                        if type == "integer":
                            default_str = 1
                        elif type == 'number':
                            default_str = 2.0
                        else:
                            default_str = "string"
                    if parameter['in'] == "header":
                        headers[parameter['name']] = str(default_str)
                    else:
                        parameters[parameter['name']] = default_str
            except Exception as pe:
                pass
        if parameters1 :
            parameters = parameters1.copy()
        if "application/json" in headers['Content-Type']:
            data = json.dumps(parameters)
        else:
            data = ""
            for key,values in parameters.items():
                data +=key+"="+str(values)+"&"
            data = data[:-1]
        new_url = url1 + path
        return new_url,headers,summary,data,method
    except Exception as e:
        return False
        # print(Style.BRIGHT + Fore.RED + "[-] 出现一小个错误！POST参数，url为：" + url + path + " , Error Message：" ,e)

def run(url,proxies,fpath,mode,cookie):
    if cookie != '':
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'content-type': 'application/json',
        }
    else:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36',
            'content-type': 'application/json',
            'Cookie': cookie
        }
    send_data_list.append(['URL','接口描述','状态码','响应包大小','请求包','响应包'])
    uploads = []
    downloads = []
    all_list = []
    user_list = []
    try:
        rep = requests.get(url=url, verify=False, headers=headers).json()
    except JSONDecodeError:
        print(Style.BRIGHT + Fore.RED + "[-] 不是有效的json接口。")
        exit()
    except Exception as e:
        print(Style.BRIGHT + Fore.RED + "[-] 程序出错了,异常信息如下：")
        print(e)
        exit()
    url1 = re.findall('^http[s]?://.+\.[0-9a-zA-Z-]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?[/]',url)[0]
    ret = re.search('^http[s]?://(?P<host>.+\.[0-9a-zA-Z-]+[:]?[1-6]?[0-9]?[0-9]?[0-9]?[0-9]?)[/]', url)
    host = ret.group('host')
    filename = host.replace(':','-')
    url1 = url1.rstrip('/')
    if fpath != "":
        url1 = url1 + '/' + fpath
    for path in rep['paths']:
        flag = screen(path)
        # if (flag == []) and (mode != "all"):
        #     continue
        for method in rep['paths'][path]:
            # print(path,method)
            if method == "get":
                get_methods = get_method(rep, path, method, url1,cookie)
                if get_methods != False:
                    all_list.append(get_methods)
                    if 'get' in flag or (mode == "all"):
                        # 防止乱带参数而没有查询到数据，直接访问接口反而有数据的情况，不过要有brup代理才看得到。
                        requests.get(url1 + path,headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36'} ,proxies=proxies, verify=False)
                        Scanner(get_methods[0], get_methods[1], get_methods[4], proxies, get_methods[2], url1 + path)
                    if 'upload' in flag:
                        uploads.append(get_methods)
                    if 'download' in flag:
                        downloads.append(get_methods)
                    if 'user' in flag:
                        user_list.append(get_methods)
            else:
                post_methods = post_method(rep, path, method, url1,cookie)
                if post_methods != False:
                    all_list.append(post_methods)
                    if ('get' in flag) or (mode == "all"):
                        # 防止乱带参数而没有查询到数据，直接访问接口反而有数据的情况，不过要有brup代理才看得到。
                        if method == "post":
                            requests.post(url1 + path,headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.82 Safari/537.36'}, proxies=proxies, verify=False)
                        Scanner(post_methods[0], post_methods[1], post_methods[4], proxies, post_methods[2], post_methods[3])
                    if 'upload' in flag:
                        uploads.append(post_methods)
                    if 'download' in flag:
                        downloads.append(post_methods)
                    if 'user' in flag:
                        user_list.append(post_methods)
            # else:
            #     print(Style.BRIGHT + Fore.RED + "[-] 请求方法既不是GET也不是POST！")
    # 创建一个工作簿
    wb = Workbook()
    # 选择要写入数据的工作表
    ws = wb.active
    # 修改第一个表的标题
    ws.title = "自动化测试的查询接口"
    # 写入数据
    for input_data in send_data_list:
        ws.append(input_data)
    if uploads != []:
        sheet1 = wb.create_sheet(title="疑似上传接口")
        sheet1.append(['!!!注意！文件上传构造的数据包并不准确，需要手动构造。!!!'])
        sheet1.append(['接口','接口描述','构造的请求包'])
        tmp_data = print_api(uploads)
        for tmp in tmp_data:
            sheet1.append(tmp)
    if downloads != []:
        sheet2 = wb.create_sheet(title="疑似下载接口")
        sheet2.append(['接口', '接口描述', '构造的请求包'])
        tmp_data = print_api(downloads)
        for tmp in tmp_data:
            sheet2.append(tmp)
    if user_list != []:
        sheet3 = wb.create_sheet(title="疑似用户相关接口")
        sheet3.append(['接口', '接口描述', '构造的请求包'])
        tmp_data = print_api(user_list)
        for tmp in tmp_data:
            sheet3.append(tmp)
    if all_list != []:
        sheet4 = wb.create_sheet(title="所有接口")
        sheet4.append(['接口', '接口描述', '构造的请求包'])
        tmp_data = print_api(all_list)
        for tmp in tmp_data:
            sheet4.append(tmp)
    # 保存
    wb.save(filename=filename+".xlsx")
    # 关闭表格
    wb.close()
    print(Style.BRIGHT + Fore.GREEN + '\n[+] 所有结果已保存到'+filename+".xlsx文件中。")

if __name__ == '__main__':
    init(autoreset=True)
    args = args()
    url = args.url if args.url[-1]!='/' else args.url[:-1]
    proxies = {'http': args.proxy, 'https': args.proxy}
    run(url,proxies,args.path,args.mode,args.cookie)